"""
Build the final QDArchive project classification deliverables in one command.

Steps:
    1. Validate the selected production classifications (cross-model rule,
       no duplicates, no invalid codes, no local-dry-run/model_error results).
    2. Generate the XLSX classification table.
    3. Generate the PDF classification report.
    4. Reopen both output files independently (not just relying on each
       generator's own internal validation) and verify they are well-formed.
    5. Print a PASS/FAIL summary and the exact output paths and file sizes.

Read-only against the database; the only files written are the XLSX, the PDF,
and each generator's own validation CSV. No API calls.

Usage:
    python phase_2/build_final_deliverables.py --db 23727550-sq26-combined.db
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

_here = Path(__file__).parent
if str(_here) not in sys.path:
    sys.path.insert(0, str(_here))

from openpyxl import load_workbook
from pypdf import PdfReader

import export_project_classification_table as xlsx_exporter
import generate_project_classification_report as pdf_reporter
from project_classification_data import (
    DEFAULT_FALLBACK_METHOD,
    DEFAULT_PREFERRED_METHOD,
    connect_readonly,
    coverage_counts,
    fetch_project_rows,
    load_isic_titles,
    print_schema_decisions,
)

DB_DEFAULT = "23727550-sq26-combined.db"
XLSX_OUTPUT_DEFAULT = xlsx_exporter.OUTPUT_DEFAULT
PDF_OUTPUT_DEFAULT = pdf_reporter.OUTPUT_DEFAULT


def _print_checks(checks: list[dict]) -> None:
    for c in checks:
        detail = f" ({c['detail']})" if c["detail"] else ""
        print(f"  [{c['status']}] {c['check']}{detail}")


def validate_selection(conn, preferred_method: str, fallback_method: str) -> list[dict]:
    """Sanity-check the cross-model classification selection itself, before
    either deliverable is generated from it."""
    checks: list[dict] = []

    def add(name: str, passed: bool, detail: str = "") -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    titles = load_isic_titles(conn)
    rows = fetch_project_rows(
        conn, preferred_method=preferred_method, fallback_method=fallback_method, include_unclassified=True,
    )
    total, covered, _remaining = coverage_counts(conn, (preferred_method, fallback_method))

    add("selection covers all eligible projects", len(rows) == total, f"{len(rows):,} rows vs {total:,} eligible")

    ids = [r["global_project_id"] for r in rows]
    dup = len(ids) - len(set(ids))
    add("no duplicate project rows in selection", dup == 0, f"{dup} duplicates")

    disallowed = {
        r["method_used"] for r in rows
        if r["method_used"] is not None and r["method_used"] not in (preferred_method, fallback_method)
    }
    add(
        "only preferred/fallback methods used",
        len(disallowed) == 0,
        f"unexpected methods: {sorted(disallowed)}" if disallowed else "none",
    )

    invalid_codes = set()
    for r in rows:
        for code in (r["primary_class_code"], r["secondary_class_code"]):
            if code and code not in titles:
                invalid_codes.add(code)
    add(
        "all selected classes map to isic_divisions",
        len(invalid_codes) == 0,
        f"unmapped: {sorted(invalid_codes)}" if invalid_codes else "none",
    )

    missing_repo = sum(1 for r in rows if r["repository_id"] is None)
    add("all rows have a repository_id", missing_repo == 0, f"{missing_repo} missing")

    classified_in_selection = sum(1 for r in rows if r["method_used"] is not None)
    add(
        "classified count matches coverage_counts",
        classified_in_selection == covered,
        f"{classified_in_selection:,} vs {covered:,}",
    )

    return checks


def reopen_verify_xlsx(path: Path, expected_data_rows: int) -> list[dict]:
    """Independent re-check: open the XLSX file that actually landed on disk
    (not the in-memory rows the exporter used) and confirm it looks right."""
    checks: list[dict] = []

    def add(name: str, passed: bool, detail: str = "") -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    if not path.exists():
        add("XLSX file exists", False, str(path))
        return checks
    add("XLSX file exists", True, f"{path.stat().st_size:,} bytes")

    wb = load_workbook(path, read_only=True)
    add("XLSX has exactly one worksheet", len(wb.sheetnames) == 1, str(wb.sheetnames))
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    add("XLSX header matches expected columns", header == xlsx_exporter.HEADERS, str(header))
    data_rows = ws.max_row - 1
    add(
        "XLSX row count matches generator output",
        data_rows == expected_data_rows,
        f"{data_rows:,} in file vs {expected_data_rows:,} generated",
    )
    wb.close()
    return checks


def reopen_verify_pdf(path: Path, expected_pages: int) -> list[dict]:
    """Independent re-check: open the PDF file that actually landed on disk
    and confirm its page count matches what was generated."""
    checks: list[dict] = []

    def add(name: str, passed: bool, detail: str = "") -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    if not path.exists():
        add("PDF file exists", False, str(path))
        return checks
    add("PDF file exists", True, f"{path.stat().st_size:,} bytes")

    reader = PdfReader(str(path))
    actual_pages = len(reader.pages)
    add(
        "PDF reopens and page count matches",
        actual_pages == expected_pages,
        f"{actual_pages} in file vs {expected_pages} generated",
    )
    return checks


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the final project classification deliverables (XLSX + PDF).")
    parser.add_argument("--db", default=DB_DEFAULT)
    parser.add_argument("--xlsx-output", default=XLSX_OUTPUT_DEFAULT)
    parser.add_argument("--pdf-output", default=PDF_OUTPUT_DEFAULT)
    parser.add_argument("--include-unclassified", action="store_true",
                         help="also include eligible-but-unclassified projects in the XLSX")
    parser.add_argument("--preferred-method", default=DEFAULT_PREFERRED_METHOD)
    parser.add_argument("--fallback-method", default=DEFAULT_FALLBACK_METHOD)
    parser.add_argument("--top-n", type=int, default=pdf_reporter.TOP_N_DEFAULT)
    args = parser.parse_args()

    all_checks: list[dict] = []

    print("=" * 64)
    print("Step 1/4: Validate selected production classifications")
    print("=" * 64)
    conn = connect_readonly(args.db)
    print_schema_decisions(args.preferred_method, args.fallback_method, args.include_unclassified)
    selection_checks = validate_selection(conn, args.preferred_method, args.fallback_method)
    conn.close()
    print()
    _print_checks(selection_checks)
    all_checks.extend(selection_checks)

    print()
    print("=" * 64)
    print("Step 2/4: Generate XLSX classification table")
    print("=" * 64)
    xlsx_rows, xlsx_checks = xlsx_exporter.run(
        db_path=args.db, output_path=args.xlsx_output, include_unclassified=args.include_unclassified,
        preferred_method=args.preferred_method, fallback_method=args.fallback_method,
    )
    all_checks.extend(xlsx_checks)

    print()
    print("=" * 64)
    print("Step 3/4: Generate PDF classification report")
    print("=" * 64)
    pdf_result = pdf_reporter.generate_report(
        db_path=args.db, output_path=args.pdf_output,
        preferred_method=args.preferred_method, fallback_method=args.fallback_method, top_n=args.top_n,
    )
    print(f"\nPDF written to {pdf_result['output_path']} ({pdf_result['page_count']} pages, "
          f"{len(pdf_result['repo_ids_processed'])} repository sections)")
    pdf_checks = pdf_reporter.validate(pdf_result, set(pdf_result["titles"].keys()))
    pdf_reporter.write_validation_report(pdf_checks)
    print()
    _print_checks(pdf_checks)
    all_checks.extend(pdf_checks)

    print()
    print("=" * 64)
    print("Step 4/4: Reopen and independently verify both outputs")
    print("=" * 64)
    xlsx_path = Path(args.xlsx_output)
    pdf_path = Path(args.pdf_output)
    reopen_xlsx_checks = reopen_verify_xlsx(xlsx_path, len(xlsx_rows))
    reopen_pdf_checks = reopen_verify_pdf(pdf_path, pdf_result["page_count"])
    _print_checks(reopen_xlsx_checks)
    _print_checks(reopen_pdf_checks)
    all_checks.extend(reopen_xlsx_checks)
    all_checks.extend(reopen_pdf_checks)

    print()
    print("=" * 64)
    print("Summary")
    print("=" * 64)
    failed = [c for c in all_checks if c["status"] != "PASS"]
    overall = "PASS" if not failed else "FAIL"
    print(f"  Overall: {overall}")
    print(f"  Total checks: {len(all_checks)}   Failed: {len(failed)}")
    print()
    print("  Output files:")
    for label, path in (("XLSX", xlsx_path), ("PDF", pdf_path)):
        size = path.stat().st_size if path.exists() else 0
        print(f"    {label}: {path}  ({size:,} bytes)")
    print("=" * 64)

    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
