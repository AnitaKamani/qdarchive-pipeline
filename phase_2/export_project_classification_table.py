"""
Export the final project classification table as XLSX.

Read-only: opens the database via a SQLite read-only URI, makes no API calls,
and does not touch Phase 1, the classification pipeline, or the schema. Uses
the cross-model selection rule in project_classification_data.py: for each
eligible project (QDA_PROJECT / QD_PROJECT), prefer --preferred-method if it
succeeded, otherwise fall back to --fallback-method, otherwise the project is
unclassified. local-dry-run and model_error rows are never considered.

By default only classified projects are written (a "classification table" is
a table of classifications). Pass --include-unclassified to also include
eligible projects with no successful classification yet, with empty
primary_class/secondary_class — the exact count of those is always reported
by generate_project_classification_report.py regardless of this flag.

Usage:
    python phase_2/export_project_classification_table.py [options]

Options:
    --db                    PATH   default: 23727550-sq26-combined.db
    --output                PATH   default: reports/23727550-sq26-project-classification-table.xlsx
    --include-unclassified         also include eligible projects with no successful
                                   classification yet (empty primary_class/secondary_class)
    --preferred-method      METHOD default: openai:gpt-4.1-mini
    --fallback-method       METHOD default: openai:gpt-4o-mini
"""

from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path

_here = Path(__file__).parent
if str(_here) not in sys.path:
    sys.path.insert(0, str(_here))

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from project_classification_data import (
    DEFAULT_FALLBACK_METHOD,
    DEFAULT_PREFERRED_METHOD,
    IGNORED_METHODS,
    connect_readonly,
    fetch_project_rows,
    isic_label,
    load_isic_titles,
    print_schema_decisions,
)

DB_DEFAULT = "23727550-sq26-combined.db"
OUTPUT_DEFAULT = "reports/23727550-sq26-project-classification-table.xlsx"
VALIDATION_REPORT = "reports/project_classification_table_validation.csv"
SHEET_NAME = "Project classifications"

HEADERS = ["repository_id", "project_type", "project_title", "primary_class", "secondary_class", "no_project_files"]

# (openpyxl column key, max width cap)
COLUMN_WIDTH_CAPS = {
    "repository_id": 16,
    "project_type": 16,
    "project_title": 60,
    "primary_class": 55,
    "secondary_class": 55,
    "no_project_files": 16,
}


def build_rows(conn, preferred_method: str, fallback_method: str, include_unclassified: bool) -> list[dict]:
    titles = load_isic_titles(conn)
    project_rows = fetch_project_rows(
        conn, preferred_method=preferred_method, fallback_method=fallback_method,
        include_unclassified=include_unclassified,
    )

    out = []
    for row in project_rows:
        out.append({
            "global_project_id": row["global_project_id"],  # kept for validation only, not written to the sheet
            "method_used": row["method_used"],               # kept for validation only, not written to the sheet
            "repository_id": row["repository_id"],
            "project_type": row["project_type"],
            "project_title": row["project_title"],
            "primary_class": isic_label(row["primary_class_code"], titles),
            "secondary_class": isic_label(row["secondary_class_code"], titles),
            "no_project_files": row["no_project_files"],
        })
    return out


def write_workbook(rows: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row[h] for h in HEADERS])

    last_col = get_column_letter(len(HEADERS))
    last_row = len(rows) + 1
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    for idx, header in enumerate(HEADERS, start=1):
        cap = COLUMN_WIDTH_CAPS.get(header, 20)
        longest = len(header)
        for row in rows:
            value = row[header]
            longest = max(longest, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(idx)].width = min(longest + 2, cap)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Validation
# ---------------------------------------------------------------------------

def validate(rows: list[dict], valid_codes: set[str], preferred_method: str, fallback_method: str,
             output_path: Path) -> list[dict]:
    checks: list[dict] = []

    def add(name: str, passed: bool, detail: str = "") -> None:
        checks.append({"check": name, "status": "PASS" if passed else "FAIL", "detail": detail})

    add(
        "exact required headers",
        HEADERS == ["repository_id", "project_type", "project_title", "primary_class", "secondary_class", "no_project_files"],
        ", ".join(HEADERS),
    )

    project_ids = [r["global_project_id"] for r in rows]
    duplicate_count = len(project_ids) - len(set(project_ids))
    add("no duplicate project rows", duplicate_count == 0, f"{duplicate_count} duplicate global_project_id values")

    missing_repo = sum(1 for r in rows if r["repository_id"] is None or r["repository_id"] == "")
    add("all repository IDs present", missing_repo == 0, f"{missing_repo} rows with missing repository_id")

    negative_files = sum(1 for r in rows if r["no_project_files"] < 0)
    add("file counts are non-negative", negative_files == 0, f"{negative_files} rows with negative no_project_files")

    def _code_from_label(label: str) -> str | None:
        if not label:
            return None
        return label.split(" — ", 1)[0]

    invalid_codes = 0
    for r in rows:
        for field in ("primary_class", "secondary_class"):
            code = _code_from_label(r[field])
            if code and code not in valid_codes:
                invalid_codes += 1
    add("all non-empty ISIC classes map to isic_divisions", invalid_codes == 0, f"{invalid_codes} unmapped class values")

    disallowed_methods = {
        r["method_used"] for r in rows if r["method_used"] is not None
    } - {preferred_method, fallback_method}
    disallowed_methods |= {r["method_used"] for r in rows if r["method_used"] in IGNORED_METHODS}
    add(
        "no local-dry-run/model_error results used",
        len(disallowed_methods) == 0,
        f"unexpected methods: {sorted(disallowed_methods)}" if disallowed_methods else "none",
    )

    # Re-open the written file to confirm it matches what we intended to write.
    if output_path.exists():
        wb = load_workbook(output_path, read_only=True)
        ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else None
        add("workbook has the expected sheet name", ws is not None, SHEET_NAME)
        if ws is not None:
            header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            add("workbook header row matches HEADERS", header_row == HEADERS, str(header_row))
            data_row_count = ws.max_row - 1
            add(
                "workbook row count matches generated rows",
                data_row_count == len(rows),
                f"{data_row_count} data rows in file vs {len(rows)} generated",
            )
        wb.close()
    else:
        add("workbook file exists", False, str(output_path))

    return checks


def write_validation_report(checks: list[dict], path: str | Path = VALIDATION_REPORT) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["check", "status", "detail"])
        w.writeheader()
        w.writerows(checks)


def run(
    db_path: str,
    output_path: str,
    include_unclassified: bool,
    preferred_method: str,
    fallback_method: str,
    validation_report_path: str | Path | None = None,
) -> tuple[list[dict], list[dict]]:
    conn = connect_readonly(db_path)
    print_schema_decisions(preferred_method, fallback_method, include_unclassified)

    titles = load_isic_titles(conn)
    rows = build_rows(conn, preferred_method, fallback_method, include_unclassified)
    print(f"\n{len(rows):,} rows selected for export.")

    out_path = Path(output_path)
    write_workbook(rows, out_path)
    print(f"Workbook written to {out_path}")

    checks = validate(rows, set(titles.keys()), preferred_method, fallback_method, out_path)
    write_validation_report(checks, validation_report_path if validation_report_path is not None else VALIDATION_REPORT)
    conn.close()

    return rows, checks


def main() -> None:
    parser = argparse.ArgumentParser(description="Export the final project classification table as XLSX.")
    parser.add_argument("--db", default=DB_DEFAULT)
    parser.add_argument("--output", default=OUTPUT_DEFAULT)
    parser.add_argument("--include-unclassified", action="store_true")
    parser.add_argument("--preferred-method", default=DEFAULT_PREFERRED_METHOD)
    parser.add_argument("--fallback-method", default=DEFAULT_FALLBACK_METHOD)
    args = parser.parse_args()

    print("=" * 64)
    print("Project Classification Table Export")
    print("=" * 64)

    rows, checks = run(
        db_path=args.db,
        output_path=args.output,
        include_unclassified=args.include_unclassified,
        preferred_method=args.preferred_method,
        fallback_method=args.fallback_method,
    )

    print()
    print("Validation:")
    all_pass = True
    for c in checks:
        detail = f" ({c['detail']})" if c["detail"] else ""
        print(f"  [{c['status']}] {c['check']}{detail}")
        if c["status"] != "PASS":
            all_pass = False
    print(f"\n  Report: {VALIDATION_REPORT}")
    print("=" * 64)

    if not all_pass:
        sys.exit(1)


if __name__ == "__main__":
    main()
